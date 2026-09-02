"""Convert workforce-report PDFs into formatted XLSX workbooks."""

from __future__ import annotations

import logging
import re
import unicodedata
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any, Iterable

import pandas as pd
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.utils.journal_logger import log_execution_entry

from .pdf_parser import PDFDocumentParser

SHIFT_OPTIONS = ("Nuit", "Soir", "Jour")
VALID_CODES = (
    "N", "S", "J", "TS1.5", "TSS", "TSN", "AIC", "SIC", "FL4", "FL7",
    "FL6", "FL8", "TSTD", "MON", "CHOC", "TRI", "EVAL", "URG", "ETJ",
    "ETS", "ETN", "BRAN", "ACUR", "HSCM", "DVERS", "CDJ", "FL", "HJT", "HOR12",
)
OVERTIME_CODES = {"TS1.5", "TSS", "TSN", "TSTD"}
LOGGER = logging.getLogger(__name__)

DEPARTMENT_PATTERNS = (
    ("HF Unité de Médecine - 4e étage", "4e"),
    ("HF Unité de Médecine - 7e étage", "7e"),
    ("HF Unité de Médecine - 6e étage", "6e"),
    ("HF Chirurgie court séjour", "8e"),
    ("HF Soins intensifs coronariens", "SIC"),
    ("HF Urgence", "URG"),
    ("HF Électrophysiologie", "ECG"),
    ("HF Accueil et réception", "ACUR/GDL"),
    ("CIUSSS Gestion des lits", "ACUR/GDL"),
)
DEPARTMENT_ORDER = ("4e", "7e", "6e", "8e", "SIC", "CDJ", "URG", "ECG", "ACUR/GDL")
CATEGORY_ORDER = ("Inf", "Aux", "PAB", "AA")
CATEGORY_PATTERNS = (
    ("Infirmière auxiliaire", "Aux"),
    ("Préposé aux bénéficiaire", "PAB"),
    ("Préposé aux bénéficiaires", "PAB"),
    ("PBM", "PAB"),
    ("PEM", "PAB"),
    ("Agent Adm 1-2-3-4", "AA"),
    ("AA3 sec et adm", "AA"),
    ("AAS sec et adm", "AA"),
    ("lafirmière", "Inf"),
    ("Infirmière", "Inf"),
)
AA_RATIO_GRID = {
    "Jour": {"4e": 1, "7e": 1, "6e": 1, "8e": 1, "SIC": 0, "CDJ": 1, "URG": 2, "ECG": 0, "ACUR/GDL": 2},
    "Soir": {"4e": 1, "7e": 1, "6e": 1, "8e": 1, "SIC": 0, "CDJ": 0, "URG": 2, "ECG": 0, "ACUR/GDL": 2},
    "Nuit": {"4e": 0, "7e": 0, "6e": 0, "8e": 0, "SIC": 0, "CDJ": 0, "URG": 1, "ECG": 0, "ACUR/GDL": 1},
}


class WorkforceReportError(ValueError):
    """Raised when a workforce report cannot be converted safely."""


DATE_PATTERN = re.compile(
    r"\b(?:Le\s+)?(?:lundi|mardi|mercredi|jeudi|vendredi|samedi|dimanche)?\s*"
    r"\d{1,2}(?:\s+au\s+\d{1,2})?\s+"
    r"(?:janv?(?:ier)?|févr?(?:ier)?|mars|avr(?:il)?|mai|juin|juil(?:let)?|"
    r"août|sept?(?:embre)?|oct(?:obre)?|nov(?:embre)?|déc(?:embre)?)\.?\s+\d{4}\b",
    flags=re.IGNORECASE,
)


def _date_phrases(text: str) -> list[str]:
    return [re.sub(r"\s+", " ", match.group(0)).strip() for match in DATE_PATTERN.finditer(text)]


def _normalized_text(text: str) -> str:
    return "".join(
        character for character in unicodedata.normalize("NFD", text.casefold())
        if unicodedata.category(character) != "Mn"
    )


def _date_key(date: str) -> str:
    return re.sub(r"\b(?:le\s+)?(?:lundi|mardi|mercredi|jeudi|vendredi|samedi|dimanche)\s*", "", _normalized_text(date)).strip(". ")


def extract_report_date(text: str) -> str:
    """Extract the complete date range represented in the report text."""
    dates = list(dict.fromkeys(_date_phrases(text)))
    if not dates:
        raise WorkforceReportError("Date du rapport introuvable (format attendu: Le [jour] [date] [mois] [année]).")
    return " | ".join(dates)


def _find_label(line: str, patterns: Iterable[tuple[str, str]]) -> str | None:
    folded = _normalized_text(line)
    for label, code in patterns:
        if _normalized_text(label) in folded:
            return code
    return None


def _find_department(line: str) -> str | None:
    normalized_line = _normalized_text(line)
    medicine_floor = re.search(
        r"\b[hm][fe]?\s+unite\s+de\s+medecine\s*-?\s*(\d+\s*(?:e|eme)?\s*etage)\b",
        normalized_line,
    )
    if medicine_floor:
        floor_number = re.match(r"\d+", medicine_floor.group(1))
        if floor_number:
            floor = int(floor_number.group(0))
            # Map floor to department code
            return f"{floor}e"
        return None
    department = _find_label(line, DEPARTMENT_PATTERNS)
    if department:
        return department
    if "chirurgie d'un jour" in normalized_line:
        return "CDJ"
    if "chirurgie court sejour" in normalized_line:
        return "8e"
    if "soins intensifs coronariens" in normalized_line:
        return "SIC"
    if "urgence" in normalized_line:
        return "URG"
    if "electrophysiologie" in normalized_line:
        return "ECG"
    if "accueil et reception" in normalized_line or "gestion des lits" in normalized_line:
        return "ACUR/GDL"
    if re.search(r"\b7(?:e|eme)?\s+etage\b", normalized_line):
        return "7e"
    if re.search(r"\b6(?:e|eme)?\s+etage\b", normalized_line):
        return "6e"
    if re.search(r"\bsic\b", normalized_line) and re.search(r"\bhf\b", normalized_line):
        return "SIC"
    if re.search(r"\becg\b", normalized_line) and re.search(r"\bhf\b", normalized_line):
        return "ECG"
    return None


def parse_ratio(ratio_str: object) -> tuple[int | None, int | None]:
    """Parse complete, partial, or missing target/presence ratios safely."""
    value = str(ratio_str)
    match = re.search(r"(\d+)\s*/\s*[^\d]{0,12}(\d+)", value)
    if match:
        return int(match.group(1)), int(match.group(2))
    single_digit = re.search(r"\b(\d{1,2})\b", value)
    if single_digit:
        return int(single_digit.group(1)), None
    return None, None


def _ratio_from_block(block: str) -> tuple[int | None, int | None]:
    ratio_label = re.search(r"Ratio\s*/\s*Présences", block, re.IGNORECASE)
    if ratio_label:
        ratio_text = block[ratio_label.end():ratio_label.end() + 20]
        match = re.match(
            r"\s*[:._'\"-]*\s*(\d{1,2})(?!\d)(?:\s*/\s*[^\d]{0,12}(\d+))?",
            ratio_text,
        )
        if match:
            return int(match.group(1)), int(match.group(2)) if match.group(2) else None
        return None, None
    if re.search(r"(?<!\d)\d+\s*/\s*\d+(?!\d)", block):
        return parse_ratio(block)
    return None, None


def _codes_in(text: str) -> list[str]:
    token_pattern = re.compile(r"(?<![A-Za-z0-9.])(?:" + "|".join(map(re.escape, sorted(VALID_CODES, key=len, reverse=True))) + r")(?![A-Za-z0-9.])")
    return token_pattern.findall(text.upper())


def _codes_in_code_column(block: str) -> list[str]:
    """Count only codes after the row's target/presence ratio."""
    ratio_match = re.search(r"(?<!\d)(\d+)\s*/\s*(\d+)(?!\d)", block)
    code_text = block[ratio_match.end():] if ratio_match else block
    return _codes_in(code_text)


def _codes_for_department(block: str, department: str) -> list[str]:
    codes = _codes_in_code_column(block)
    if department != "URG":
        return [code for code in codes if code != "DVERS"]
    return codes


STRUCTURAL_VALUE_PATTERN = re.compile(
    r"\b(?:\d{1,2}:\d{2}|\d{4,}|[A-Za-z]{1,3}\d{2,})\b",
    re.IGNORECASE,
)
STRUCTURAL_HEADER_PATTERN = re.compile(
    r"\b(?:employ[ée]?|te|entr[ée]e|sortie|repas|code\s+repas|no\.?\s*poste|code)\b",
    re.IGNORECASE,
)
EMPLOYMENT_ID_PATTERN = re.compile(r"\b(891\d|847\d|2490|3455|3480|531[57])\b")
POSITION_ID_PATTERN = re.compile(r"\b\d{6,7}\b")


def _category_from_staff_row(line: str) -> str | None:
    employee_id = EMPLOYMENT_ID_PATTERN.search(line)
    if not employee_id:
        return None
    value = employee_id.group(1)
    if value in {"3455"}:
        return "Aux"
    if value in {"3480"}:
        return "PAB"
    if value.startswith("531"):
        return "AA"
    return "Inf"


def _is_physical_staff_row(line: str) -> bool:
    """Return whether an extracted line contains a staff-table value, not a header."""
    if _date_phrases(line):
        return False
    return bool(STRUCTURAL_VALUE_PATTERN.search(line)) and not bool(
        STRUCTURAL_HEADER_PATTERN.search(line)
    )


def _count_physical_staff_rows(lines: Iterable[str]) -> int:
    """Count data rows from table structure without using the employment code value."""
    data_rows = [line for line in lines if _is_physical_staff_row(line)]
    employment_rows = [line for line in data_rows if EMPLOYMENT_ID_PATTERN.search(line)]
    position_rows = [
        line for line in data_rows
        if POSITION_ID_PATTERN.search(line) and not EMPLOYMENT_ID_PATTERN.search(line)
    ]
    # Some native PDF extracts separate the employee/entry and position/code columns.
    # Their rows represent the same people, so count the longer projection.
    if any(not POSITION_ID_PATTERN.search(line) for line in employment_rows):
        return max(len(employment_rows), len(position_rows))
    return len(data_rows)


def _recount_acur_gdl_aa_rows(lines: Iterable[str]) -> int:
    """Count only physical administrative rows in one ACUR/GDL AA block."""
    return sum(
        _is_physical_staff_row(line) and _category_from_staff_row(line) == "AA"
        for line in lines
    )


def _apply_acur_gdl_category_rule(
    department: str, category: str, target: int, presence: int
) -> tuple[int, int]:
    """Keep ACUR/GDL data exclusively in the AA category."""
    if department == "ACUR/GDL" and category in {"Inf", "Aux", "PAB"}:
        return 0, 0
    return target, presence


def _target_for(category: str, department: str, shift: str, explicit_target: int | None) -> int:
    """Prefer an explicit PDF target; otherwise use the AA reference grid."""
    if explicit_target is not None:
        return explicit_target
    if category == "AA":
        return AA_RATIO_GRID[shift][department]
    return 0


def _has_date_and_department_on_line(line: str) -> bool:
    """Check if a line contains both a date and a department marker."""
    has_date = bool(_date_phrases(line))
    has_department = bool(_find_department(line))
    return has_date and has_department


def _error(report_date: str, shift: str, category: str, employment_code: str, detail: str) -> WorkforceReportError:
    return WorkforceReportError(
        f"{detail} | Date: {report_date} | Quart: {shift} | Catégorie d'emploi: {category} | Code d'emploi: {employment_code}"
    )


def parse_workforce_text(
    text: str, shift: str, warnings: list[str] | None = None
) -> list[dict[str, Any]]:
    """Parse report text into validated department/category records."""
    if shift not in SHIFT_OPTIONS:
        raise ValueError(f"Quart invalide: {shift}")
    report_date = extract_report_date(text)
    source_lines = [line.strip() for line in text.splitlines() if line.strip()]
    lines: list[str] = []
    inferred_category: str | None = None
    for line in source_lines:
        department = _find_department(line)
        explicit_category = _find_label(line, CATEGORY_PATTERNS)
        if department:
            inferred_category = None
        if explicit_category:
            inferred_category = explicit_category
        staff_category = _category_from_staff_row(line)
        if staff_category and inferred_category != staff_category:
            lines.append(next(label for label, code in CATEGORY_PATTERNS if code == staff_category))
            inferred_category = staff_category
        lines.append(line)
    records: list[dict[str, Any]] = []
    current_department: str | None = None
    current_date = report_date.split(" | ")[0]
    primary_date_key = _date_key(current_date)
    sic_occurrence = 0
    for index, line in enumerate(lines):
        dates_on_line = _date_phrases(line)
        if dates_on_line:
            detected_date = dates_on_line[0]
            if _date_key(detected_date) != primary_date_key:
                current_date = detected_date
        department = _find_department(line)
        if department:
            if department == "SIC":
                sic_occurrence += 1
                current_department = "CDJ" if sic_occurrence in {3, 4} else department
            else:
                current_department = department
        category = _find_label(line, CATEGORY_PATTERNS)
        if not category:
            continue
        if current_department is None:
            # Look ahead to find department in the current block
            for following in lines[index + 1:]:
                found_dept = _find_department(following)
                if found_dept:
                    current_department = found_dept
                    break
                if _find_label(following, CATEGORY_PATTERNS):
                    break
        if current_department is None:
            raise _error(report_date, shift, category, category, "Département introuvable")

        block_lines = [line]
        for following in lines[index + 1:]:
            if _find_department(following) or _find_label(following, CATEGORY_PATTERNS):
                break
            block_lines.append(following)
        block = " ".join(block_lines)
        codes = _codes_for_department(block, current_department)
        presence = _count_physical_staff_rows(block_lines[1:])
        # Apply business rule: ignore ratios on lines that contain both date and department
        if _has_date_and_department_on_line(line):
            target, stated_presence = None, None
        else:
            target, stated_presence = _ratio_from_block(block)
        if target is not None:
            if target > 20:
                warning = (
                    f"Valeur Cible OCR improbable [{current_date} | {shift} | "
                    f"{current_department} | {category}] : {target} remplacée par 0."
                )
                LOGGER.warning(warning)
                if warnings is not None:
                    warnings.append(warning)
                target = None
            if (
                stated_presence is not None
                and stated_presence != presence
                and not (current_department == "ACUR/GDL" and category in {"Inf", "Aux", "PAB"})
            ):
                warning = (
                    f"Écart détecté [{current_date} | {shift} | {current_department} | {category}] : "
                    f"Présences indiquées = {stated_presence}, Lignes comptées = {presence}. "
                    f"La valeur {presence} a été retenue pour le fichier Excel."
                )
                LOGGER.warning(warning)
                if warnings is not None:
                    warnings.append(warning)
        if target is None and category != "AA":
            target = presence
        else:
            target = _target_for(category, current_department, shift, target)
        if current_department == "ACUR/GDL" and category == "AA" and presence > 5:
            warning = (
                "Décompte élevé détecté pour AA dans ACUR/GDL (>5). "
                "Vérification de sécurité déclenchée."
            )
            LOGGER.warning(warning)
            if warnings is not None:
                warnings.append(warning)
            presence = _recount_acur_gdl_aa_rows(block_lines[1:])
        target, presence = _apply_acur_gdl_category_rule(
            current_department, category, target, presence
        )
        extra_codes = codes[target:] if presence > target else []
        if presence > target:
            suffix = f"+{presence - target}TS" if any(code in OVERTIME_CODES for code in extra_codes) else f"+{presence - target}R"
        elif presence < target:
            suffix = f"-{target - presence}"
        else:
            suffix = ""
        records.append({"Département": current_department, "Catégorie": category, "Cible": target, "Présences": presence, "Décompte des lignes": presence, "Écart": suffix, "Codes": codes, "Date": current_date})
    if not records:
        raise WorkforceReportError(f"Aucune catégorie d'effectif trouvée | Date: {report_date} | Quart: {shift} | Catégorie d'emploi: inconnue | Code d'emploi: inconnu")
    return records


def _complete_date_skeleton(
    date_records: list[dict[str, Any]], date: str, shift: str
) -> list[dict[str, Any]]:
    present_pairs = {
        (record.get("Département", ""), record.get("Catégorie", ""))
        for record in date_records
    }
    completed = list(date_records)
    for department in DEPARTMENT_ORDER:
        for category in CATEGORY_ORDER:
            if (department, category) in present_pairs:
                continue
            completed.append(
                {
                    "Département": department,
                    "Catégorie": category,
                    "Cible": _apply_acur_gdl_category_rule(
                        department, category, _target_for(category, department, shift, None), 0
                    )[0],
                    "Présences": 0,
                    "Décompte des lignes": 0,
                    "Écart": "",
                    "Codes": [],
                    "Date": date,
                }
            )
    return completed


def build_workforce_workbook(
    records: list[dict[str, Any]], shift: str, warnings: list[str] | None = None
) -> BytesIO:
    """Create a formatted workbook from validated workforce records."""
    if not records:
        raise ValueError("Au moins une ligne d'effectif est requise")
    output = BytesIO()
    headers = ("Département", "Catégorie", "Cible", "Présences", "Écart (Décompte vs Cible)")
    records_by_date: dict[str, list[dict[str, Any]]] = {}
    for record in records:
        date = record.get("Date", "") or "Date inconnue"
        records_by_date.setdefault(date, []).append(record)
    
    # Deduplicate categories per department per date and enforce category order
    def deduplicate_and_sort_records(date_records: list[dict[str, Any]]) -> list[dict[str, Any]]:
        seen = set()
        deduplicated = []
        for record in date_records:
            dept = record.get("Département", "")
            cat = record.get("Catégorie", "")
            key = (dept, cat)
            if key not in seen:
                seen.add(key)
                deduplicated.append(record)
        
        # Sort by category order within the deduplicated list
        category_order_idx = {cat: idx for idx, cat in enumerate(CATEGORY_ORDER)}
        deduplicated.sort(key=lambda r: (
            DEPARTMENT_ORDER.index(r.get("Département", "")) if r.get("Département", "") in DEPARTMENT_ORDER else len(DEPARTMENT_ORDER),
            category_order_idx.get(r.get("Catégorie", ""), len(CATEGORY_ORDER))
        ))
        return deduplicated
    
    records_by_date = {
        date: deduplicate_and_sort_records(_complete_date_skeleton(date_records, date, shift))
        for date, date_records in records_by_date.items()
    }

    def sheet_name_for(date: str, used_names: set[str]) -> str:
        base_name = re.sub(r"[\\/*?:\[\]]", "-", date).strip() or "Effectifs"
        base_name = base_name[:31]
        name = base_name
        suffix = 2
        while name in used_names:
            suffix_text = f" ({suffix})"
            name = f"{base_name[:31 - len(suffix_text)]}{suffix_text}"
            suffix += 1
        used_names.add(name)
        return name

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        workbook = writer.book
        used_sheet_names: set[str] = set()
        sheets = []
        audit_rows = []
        hor12_formatting_rows: list[tuple[Any, int]] = []  # Track (sheet, row_number) for HOR12 formatting
        for report_date, date_records in records_by_date.items():
            rows = []
            for idx, record in enumerate(date_records):
                target, presence = _apply_acur_gdl_category_rule(
                    record.get("Département", ""),
                    record.get("Catégorie", ""),
                    record.get("Cible", 0),
                    record.get("Présences", 0),
                )
                rows.append({
                    "Département": record.get("Département", ""),
                    "Catégorie": record.get("Catégorie", ""),
                    "Cible": target,
                    "Présences": presence,
                    "_codes": record.get("Codes", []),  # Track codes for HOR12 detection
                })
            dataframe = pd.DataFrame(rows, columns=(*headers[:4], "_codes")).fillna(
                {"Cible": 0, "Présences": 0}
            )
            dataframe["Cible"] = pd.to_numeric(dataframe["Cible"], errors="coerce").fillna(0)
            dataframe["Présences"] = pd.to_numeric(
                dataframe["Présences"], errors="coerce"
            ).fillna(0)
            department_dtype = pd.api.types.CategoricalDtype(
                categories=DEPARTMENT_ORDER, ordered=True
            )
            dataframe["Département"] = dataframe["Département"].astype(department_dtype)
            # Sort by department first, then by category order
            category_order_idx = {cat: idx for idx, cat in enumerate(CATEGORY_ORDER)}
            dataframe["_category_order"] = dataframe["Catégorie"].map(lambda x: category_order_idx.get(x, len(CATEGORY_ORDER)))
            dataframe = dataframe.sort_values(
                by=["Département", "_category_order"], kind="stable", na_position="last"
            ).reset_index(drop=True)
            dataframe = dataframe.drop(columns=["_category_order"], errors="ignore")
            codes_list = dataframe.pop("_codes")
            dataframe["Écart (Décompte vs Cible)"] = dataframe["Présences"] - dataframe["Cible"]
            
            sheet_name = sheet_name_for(report_date, used_sheet_names)
            dataframe.to_excel(writer, sheet_name=sheet_name, startrow=3, index=False)
            sheet = writer.sheets[sheet_name]
            sheets.append((sheet, dataframe))
            
            # Track rows with HOR12 codes for formatting
            for idx, codes in enumerate(codes_list):
                if codes and "HOR12" in codes:
                    excel_row = idx + 5  # +5 because of header rows (startrow=3 + header row 4)
                    hor12_formatting_rows.append((sheet, excel_row, dataframe.iloc[idx]))
            
            audit_rows.append(
                {
                    "Date": report_date,
                    "Total Cible": int(dataframe["Cible"].sum()),
                    "Total Présences": int(dataframe["Présences"].sum()),
                    "Total Écart": int(dataframe["Écart (Décompte vs Cible)"].sum()),
                    "Anomalies Flagged": sum(
                        report_date in warning for warning in (warnings or [])
                    ),
                }
            )

            sheet.merge_cells("A1:E1")
            sheet["A1"] = shift
            sheet.merge_cells("A2:E2")
            sheet["A2"] = report_date
        audit = workbook.create_sheet("Rapport_Audit", 0)
        audit.merge_cells("A1:E1")
        audit["A1"] = "Rapport d'audit d'exécution"
        audit.merge_cells("A2:E2")
        audit["A2"] = f"Exécuté le : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Quart : {shift}"
        audit_headers = ("Date", "Total Cible", "Total Présences", "Total Écart", "Anomalies Flagged")
        for column, header in enumerate(audit_headers, 1):
            audit.cell(4, column, header)
        for row, values in enumerate(audit_rows, 5):
            for column, header in enumerate(audit_headers, 1):
                audit.cell(row, column, values[header])
        anomaly_header_row = max(6, 6 + len(audit_rows))
        audit.cell(anomaly_header_row, 1, "Journal des anomalies")
        for row, warning in enumerate(warnings or [], anomaly_header_row + 1):
            audit.cell(row, 1, warning)
            audit.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)

    navy = "17324D"
    pale = "E8EEF3"
    thin_gray = Side(style="thin", color="B8C2CC")
    red_fill = PatternFill("solid", fgColor="FFC7CE")
    green_fill = PatternFill("solid", fgColor="C6EFCE")
    for row in audit.iter_rows(min_row=1, max_row=max(4, audit.max_row), min_col=1, max_col=5):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in audit[1] + audit[2] + audit[4]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=navy)
    for column, width in enumerate((28, 18, 20, 16, 20), 1):
        audit.column_dimensions[get_column_letter(column)].width = width
    audit.freeze_panes = "A5"
    for sheet, dataframe in sheets:
        for cell in sheet[1] + sheet[2]:
            cell.font = Font(bold=True, size=14 if cell.row == 1 else 11, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=navy)
            cell.alignment = Alignment(horizontal="center")
        for cell in sheet[4]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=navy)
            cell.alignment = Alignment(horizontal="center")
        for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row, min_col=1, max_col=5):
            for cell in row:
                cell.border = Border(bottom=thin_gray)
                if cell.row > 4 and cell.row % 2 == 1:
                    cell.fill = PatternFill("solid", fgColor=pale)
        for column in ("E",):
            for cell in sheet[column][4:]:
                if cell.value != 0:
                    cell.fill = red_fill
            sheet.conditional_formatting.add(
                f"{column}5:{column}{sheet.max_row}",
                CellIsRule(operator="notEqual", formula=["0"], fill=red_fill),
            )
        for column, width in enumerate((20, 18, 12, 14, 24), 1):
            sheet.column_dimensions[get_column_letter(column)].width = width
        sheet.freeze_panes = "A5"
    
    # Apply HOR12 formatting (dark blue background with white text)
    dark_blue_fill = PatternFill("solid", fgColor="17324D")  # Same dark blue as headers
    white_font = Font(color="FFFFFF", bold=True)
    for sheet, excel_row, row_data in hor12_formatting_rows:
        cell = sheet.cell(excel_row, 5)
        current_value = cell.value or 0
        cell.value = f"{current_value}+HOR12"
        cell.fill = dark_blue_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    workbook.save(output)
    output.seek(0)
    return output


def convert_workforce_pdf(
    pdf_path: str | Path,
    shift: str,
    warnings: list[str] | None = None,
    source_file_name: str | None = None,
) -> BytesIO:
    """Extract a PDF report and return its formatted XLSX representation."""
    try:
        parsed = PDFDocumentParser().parse(pdf_path)
        if parsed.get("status") != "success":
            raise WorkforceReportError(parsed.get("message", "Extraction PDF impossible"))
        records = parse_workforce_text(parsed["raw_text"], shift, warnings)
        output = build_workforce_workbook(records, shift, warnings)
        log_execution_entry(
            source_file=source_file_name or str(pdf_path),
            anonymized=False,
            report_date=" | ".join(dict.fromkeys(record.get("Date", "") for record in records)),
            shift=shift,
            records=records,
            warnings=warnings,
        )
        return output
    except WorkforceReportError:
        raise
    except Exception as exc:
        raise WorkforceReportError(f"Erreur de conversion du rapport: {exc}") from exc