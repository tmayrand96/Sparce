from io import BytesIO
from openpyxl import load_workbook

from backend.core.workforce_pipeline import build_workforce_workbook, parse_workforce_text
from src.aggregator import target_for
from src.utils.journal_logger import log_execution_entry


REPORT_TEXT = """Le vendredi 4 sept. 2026
HF Urgence
Infirmière 1/1
8911 16:15 23:30 00:45 301768 URG
HF Unité de Médecine - 4e étage
Infirmière 3/4
8914 15:15 23:30 00:45 309079 N
8489 16:15 23:30 00:45 305792 S
2490 15:15 23:30 00:45 309931 J
8912 15:15 23:30 00:45 309932 TS1.5
Agent Adm 1-2-3-4
5317 15:30 23:30 01:00 309084 ACUR
5317 15:45 23:45 01:00 309085 HSCM
"""


def test_parse_report_applies_shift_exception_and_validates_codes():
    records = parse_workforce_text(REPORT_TEXT, "Soir")

    assert records[0]["Département"] == "URG"
    assert records[0]["Cible"] == 9
    assert records[0]["Présences"] == 1
    assert records[0]["Écart"] == "-8"
    assert records[1]["Écart"] == "+1TS"
    assert records[2]["Cible"] == 1
    assert records[2]["Présences"] == 2


def test_presence_counts_structured_rows_without_using_employment_codes():
    text = """Le vendredi 4 sept. 2026
HF Urgence
Infirmière Ratio/Présences 3/3
Employé TE Entrée Sortie Repas Code repas No. poste Code
8911 16:15 23:30 00:45 301768 INCONNU
16:15 23:30 00:45 304776
8914 15:15 23:30 00:45 308749 TRI
"""

    record = parse_workforce_text(text, "Soir")[0]

    assert record["Présences"] == 3
    assert record["Décompte des lignes"] == 3
    assert record["Codes"] == ["TRI"]


def test_presence_ignores_table_headers_and_dates():
    text = """Le vendredi 4 sept. 2026
HF Urgence
Infirmière Ratio/Présences 1/1
Employé TE Entrée Sortie Repas Code repas No. poste Code
Le vendredi 4 sept. 2026
8911 16:15 23:30 00:45 301768 TRI
"""

    record = parse_workforce_text(text, "Soir")[0]

    assert record["Présences"] == 1


def test_fixed_targets_ignore_ocr_ratios():
    text = """Le vendredi 4 sept. 2026
HF Unité de Médecine - 4e étage
Agent Adm 1-2-3-4
5317 15:30 23:30 01:00 309084 FL4
HF Urgence
Agent Adm 1-2-3-4
5317 15:30 23:30 01:00 309085 URG
HF Accueil et réception
Agent Adm 1-2-3-4 7/1
5317 15:30 23:30 01:00 309086 ACUR
"""

    records = parse_workforce_text(text, "Soir")

    assert [(record["Département"], record["Cible"], record["Présences"]) for record in records] == [
        ("4e", 1, 1),
        ("URG", 2, 1),
        ("ACUR/GDL", 2, 1),
    ]


def test_workbook_contains_shift_date_and_formatted_rows():
    records = parse_workforce_text(REPORT_TEXT, "Soir")
    workbook = build_workforce_workbook(records, "Soir")

    loaded_workbook = load_workbook(BytesIO(workbook.getvalue()))
    sheet = loaded_workbook["Le vendredi 4 sept. 2026"]
    assert sheet["A1"].value == "Soir"
    assert sheet["A2"].value == "Le vendredi 4 sept. 2026"
    assert [cell.value for cell in sheet[4]] == [
        "Département", "Catégorie", "Cible", "Présences", "Écart (Décompte vs Cible)"
    ]
    fourth_floor_row = next(row for row in range(5, sheet.max_row + 1) if sheet.cell(row, 1).value == "4e")
    assert sheet.cell(fourth_floor_row, 5).value == 1
    assert sheet.cell(fourth_floor_row, 5).fill.fgColor.rgb == "00FFC7CE"
    assert sheet.freeze_panes == "A5"


def test_presence_uses_counted_rows_without_reading_ocr_ratio():
    text = "Le vendredi 4 sept. 2026\nHF Urgence\nInfirmière 2/1\n8911 16:15 23:30 00:45 301768 URG\n8471 16:15 23:30 00:45 304776 N"

    warnings = []
    records = parse_workforce_text(text, "Nuit", warnings)
    workbook = build_workforce_workbook(records, "Nuit")
    sheet = load_workbook(BytesIO(workbook.getvalue()))["Le vendredi 4 sept. 2026"]

    assert records[0]["Présences"] == 2
    urg_row = next(row for row in range(5, sheet.max_row + 1) if sheet.cell(row, 1).value == "URG")
    assert sheet.cell(urg_row, 4).value == 2
    assert warnings == []


def test_anonymized_report_keeps_rows_with_flexible_department_header():
    text = """Le vendredi 4 sept. 2026
HF Unité de Médecine 4e étage
Infirmière Ratio/Présences 1
15:15 23:30 00:45 309079 N
"""

    records = parse_workforce_text(text, "Nuit")

    assert records[0]["Département"] == "4e"
    assert records[0]["Présences"] == 1
    assert records[0]["Codes"] == ["N"]


def test_workbook_defaults_missing_target_and_presence_to_zero():
    workbook = build_workforce_workbook(
        [{"Département": "URG", "Catégorie": "Inf", "Date": "Le 4 sept. 2026"}],
        "Nuit",
    )

    sheet = load_workbook(BytesIO(workbook.getvalue()))["Le 4 sept. 2026"]

    urg_inf_row = next(
        row for row in range(5, sheet.max_row + 1)
        if sheet.cell(row, 1).value == "URG" and sheet.cell(row, 2).value == "Inf"
    )
    assert sheet.cell(urg_inf_row, 3).value == 7
    assert sheet.cell(urg_inf_row, 4).value == 0
    assert sheet.cell(urg_inf_row, 5).value == -7
    assert sheet.cell(urg_inf_row, 5).fill.fgColor.rgb == "00FFC7CE"


def test_workbook_uses_one_sheet_per_report_date():
    text = """Le lundi 7 sept. 2026
HF Urgence
Infirmière 1/1 N
Le mardi 8 sept. 2026
HF Urgence
Infirmière 1/1 J
"""

    records = parse_workforce_text(text, "Nuit")
    workbook = load_workbook(BytesIO(build_workforce_workbook(records, "Nuit").getvalue()))

    assert workbook.sheetnames == ["Rapport_Audit", "Le lundi 7 sept. 2026", "Le mardi 8 sept. 2026"]
    assert workbook[workbook.sheetnames[1]]["A2"].value == "Le lundi 7 sept. 2026"
    assert workbook[workbook.sheetnames[2]]["A2"].value == "Le mardi 8 sept. 2026"


def test_workbook_sorts_departments_in_business_order():
    records = [
        {"Département": department, "Catégorie": "Inf", "Cible": 1, "Présences": 1, "Date": "Le 7 sept. 2026"}
        for department in ("URG", "4e", "ACUR/GDL", "CDJ", "7e", "ECG", "8e", "SIC", "6e")
    ]

    workbook = load_workbook(BytesIO(build_workforce_workbook(records, "Nuit").getvalue()))
    sheet = workbook["Le 7 sept. 2026"]
    departments = list(dict.fromkeys(sheet.cell(row, 1).value for row in range(5, sheet.max_row + 1)))

    assert departments == ["4e", "7e", "6e", "8e", "SIC", "CDJ", "URG", "ECG", "ACUR/GDL"]


def test_each_date_sheet_contains_complete_department_category_skeleton():
    records = [{"Département": "URG", "Catégorie": "Inf", "Cible": 2, "Présences": 1, "Date": "Le 7 sept. 2026"}]

    workbook = load_workbook(BytesIO(build_workforce_workbook(records, "Nuit").getvalue()))
    sheet = workbook["Le 7 sept. 2026"]
    pairs = {(sheet.cell(row, 1).value, sheet.cell(row, 2).value) for row in range(5, sheet.max_row + 1)}

    assert len(pairs) == 36
    assert all((department, category) in pairs for department in ("4e", "7e", "6e", "8e", "SIC", "CDJ", "URG", "ECG", "ACUR/GDL") for category in ("Inf", "Aux", "PAB", "AA"))


def test_audit_tab_records_execution_summary_and_ocr_flags():
    warnings = []
    records = parse_workforce_text(
        "Le lundi 7 sept. 2026\nHF Urgence\nInfirmière 25/1 URG N",
        "Nuit",
        warnings,
    )

    workbook = load_workbook(BytesIO(build_workforce_workbook(records, "Nuit", warnings).getvalue()))
    audit = workbook["Rapport_Audit"]

    assert workbook.sheetnames[0] == "Rapport_Audit"
    assert audit["A1"].value == "Rapport d'audit d'exécution"
    assert audit["A5"].value == "Le lundi 7 sept. 2026"
    assert audit["B5"].value == 37
    assert audit["E5"].value == 0
    assert audit["A7"].value == "Journal des anomalies"


def test_business_rules_remap_sic_count_dvers_and_apply_aa_night_target():
    text = """Le lundi 7 sept. 2026
HF Soins intensifs coronariens
Infirmière 1/1
8911 16:15 23:30 00:45 301768 SIC
HF Soins intensifs coronariens
Infirmière auxiliaire 1/1
3455 16:15 23:30 00:45 301769 SIC
HF Soins intensifs coronariens
Infirmière 1/1
8912 16:15 23:30 00:45 301770 SIC
HF Soins intensifs coronariens
Infirmière auxiliaire 1/1
3455 16:15 23:30 00:45 301771 SIC
HF Urgence
Infirmière auxiliaire 2/2
3455 16:15 23:30 00:45 301772 DVERS
HF Accueil et réception
Agent Adm 1-2-3-4
5317 15:30 23:30 01:00 309084 ACUR
"""

    records = parse_workforce_text(text, "Nuit")
    sic_records = [record for record in records if record["Codes"] == ["SIC"]]
    urg_aux = next(record for record in records if record["Département"] == "URG" and record["Catégorie"] == "Aux")
    workbook = load_workbook(BytesIO(build_workforce_workbook(records, "Nuit").getvalue()))
    sheet = workbook["Le lundi 7 sept. 2026"]

    assert [record["Département"] for record in sic_records] == ["SIC", "SIC", "CDJ", "CDJ"]
    assert urg_aux["Présences"] == 1
    assert urg_aux["Cible"] == 1
    acur_row = next(
        row for row in range(5, sheet.max_row + 1)
        if sheet.cell(row, 1).value == "ACUR/GDL" and sheet.cell(row, 2).value == "AA"
    )
    assert sheet.cell(acur_row, 3).value == 1
    assert sheet.cell(acur_row, 4).value == 1
    assert sheet.cell(acur_row, 5).value == 0


def test_acur_gdl_keeps_only_aa_and_recounts_high_presence():
    aa_rows = "\n".join(
        f"5317 15:30 23:30 01:00 3090{index:02d} ACUR"
        for index in range(1, 7)
    )
    text = f"""Le lundi 7 sept. 2026
HF Accueil et réception
Infirmière 4/4
8911 15:30 23:30 01:00 301768 N
Infirmière auxiliaire 3/3
3455 15:30 23:30 01:00 301769 N
Préposé aux bénéficiaires 2/2
3480 15:30 23:30 01:00 301770 N
Agent Adm 1-2-3-4 2/9
{aa_rows}
"""

    warnings = []
    records = parse_workforce_text(text, "Soir", warnings)
    acur_records = {record["Catégorie"]: record for record in records if record["Département"] == "ACUR/GDL"}
    workbook = load_workbook(BytesIO(build_workforce_workbook(records, "Soir", warnings).getvalue()))
    sheet = workbook["Le lundi 7 sept. 2026"]

    assert [(acur_records[category]["Cible"], acur_records[category]["Présences"])
            for category in ("Inf", "Aux", "PAB")] == [(0, 0), (0, 0), (0, 0)]
    assert acur_records["AA"]["Cible"] == 2
    assert acur_records["AA"]["Présences"] == 6
    assert "Décompte élevé détecté pour AA dans ACUR/GDL (>5). Vérification de sécurité déclenchée." in warnings
    for category in ("Inf", "Aux", "PAB"):
        row = next(
            row for row in range(5, sheet.max_row + 1)
            if sheet.cell(row, 1).value == "ACUR/GDL" and sheet.cell(row, 2).value == category
        )
        assert (sheet.cell(row, 3).value, sheet.cell(row, 4).value) == (0, 0)


def test_high_acur_gdl_aa_warning_is_written_to_execution_journal(tmp_path):
    warning = "Décompte élevé détecté pour AA dans ACUR/GDL (>5). Vérification de sécurité déclenchée."

    journal = log_execution_entry(
        source_file="Rapport-Template-Soir.pdf",
        anonymized=False,
        report_date="Le lundi 7 sept. 2026",
        shift="Soir",
        records=[],
        warnings=[warning],
        journal_path=tmp_path / "task_journal.md",
    )

    assert "### ⚠️ Execution Warnings\n" + warning in journal


def test_workbook_evening_targets_match_reference_matrix():
    workbook = load_workbook(
        BytesIO(
            build_workforce_workbook(
                [{"Département": "URG", "Catégorie": "Inf", "Date": "Le 7 sept. 2026"}],
                "Soir",
            ).getvalue()
        )
    )
    sheet = workbook["Le 7 sept. 2026"]
    actual = {
        (sheet.cell(row, 1).value, sheet.cell(row, 2).value): sheet.cell(row, 3).value
        for row in range(5, sheet.max_row + 1)
    }
    expected = {
        "4e": (3, 2, 2, 1), "7e": (3, 1, 2, 1), "6e": (3, 2, 2, 1),
        "8e": (3, 2, 2, 1), "SIC": (4, 0, 1, 0), "CDJ": (1, 1, 0, 0),
        "URG": (9, 1, 3, 2), "ECG": (0, 0, 1, 0), "ACUR/GDL": (0, 0, 0, 2),
    }

    assert actual == {
        (department, category): targets[index]
        for department, targets in expected.items()
        for index, category in enumerate(("Inf", "Aux", "PAB", "AA"))
    }


def test_cdj_targets_are_zero_for_each_category_on_weekends():
    records = parse_workforce_text(
        """Le samedi 5 sept. 2026
HF Soins intensifs coronariens
Infirmière 1/1 SIC
HF Soins intensifs coronariens
Infirmière auxiliaire 1/1 SIC
HF Soins intensifs coronariens
Préposé aux bénéficiaires 1/1 SIC
HF Soins intensifs coronariens
Agent Adm 1-2-3-4 1/1 SIC
""",
        "Soir",
    )

    cdj_records = [record for record in records if record["Département"] == "CDJ"]
    workbook = load_workbook(BytesIO(build_workforce_workbook(records, "Soir").getvalue()))
    sheet = workbook["Le samedi 5 sept. 2026"]
    cdj_targets = [
        sheet.cell(row, 3).value
        for row in range(5, sheet.max_row + 1)
        if sheet.cell(row, 1).value == "CDJ"
    ]

    assert [record["Cible"] for record in cdj_records] == [0, 0]
    assert cdj_targets == [0, 0, 0, 0]
    assert target_for("Inf", "CDJ", "Soir", "5 sept. 2026") == 0


def test_sixth_floor_anchor_isolated_from_adjacent_department():
    text = """Le lundi 7 sept. 2026
HF Unité de médecine   6e
Infirmière
8911 15:30 23:30 01:00 301768 N
Infirmière auxiliaire
3455 15:30 23:30 01:00 301769 S
Préposé aux bénéficiaires
3480 15:30 23:30 01:00 301770 J
Agent Adm 1-2-3-4
5317 15:30 23:30 01:00 301771 FL6
HF Chirurgie court séjour
Infirmière
8912 15:30 23:30 01:00 301772 FL8
"""

    records = parse_workforce_text(text, "Soir")
    sixth_floor_records = [record for record in records if record["Département"] == "6e"]

    assert [(record["Catégorie"], record["Cible"], record["Présences"]) for record in sixth_floor_records] == [
        ("Inf", 3, 1), ("Aux", 2, 1), ("PAB", 2, 1), ("AA", 1, 1)
    ]
    assert [(record["Département"], record["Catégorie"]) for record in records[-1:]] == [("8e", "Inf")]


def test_hor12_is_excluded_from_presences_without_difference_suffix():
    text = """Le lundi 7 sept. 2026
HF Urgence
Infirmière
8911 15:30 23:30 01:00 301768 N
8912 15:30 23:30 01:00 301769 HOR12
"""

    records = parse_workforce_text(text, "Soir")
    record = records[0]
    sheet = load_workbook(BytesIO(build_workforce_workbook(records, "Soir").getvalue()))[
        "Le lundi 7 sept. 2026"
    ]
    urg_inf_row = next(
        row for row in range(5, sheet.max_row + 1)
        if sheet.cell(row, 1).value == "URG" and sheet.cell(row, 2).value == "Inf"
    )

    assert (record["Cible"], record["Présences"], record["Écart"]) == (9, 1, "-8")
    assert sheet.cell(urg_inf_row, 5).value == -8