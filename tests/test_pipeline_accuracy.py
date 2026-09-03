"""End-to-end accuracy harness for versioned workforce-report gold standards."""

from __future__ import annotations

import re
from collections import Counter
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import load_workbook

from backend.core.pdf_parser import PDFDocumentParser
from backend.core.workforce_pipeline import build_workforce_workbook, parse_workforce_text


TESTS_DIRECTORY = Path(__file__).parent
PDF_PATTERN = re.compile(r"^Rapport-A-(?P<day>\d{1,2}sept)-(?P<shift>[^.]+)\.pdf$", re.IGNORECASE)
COLUMNS_TO_COMPARE = ("Cible", "Présences", "Écart (Décompte vs Cible)")


@dataclass(frozen=True)
class GoldStandardPair:
    pdf_path: Path
    reference_path: Path
    day: str
    shift: str


def _gold_standard_pairs() -> list[GoldStandardPair]:
    pairs = []
    for pdf_path in sorted(TESTS_DIRECTORY.glob("Rapport-A-*.pdf")):
        match = PDF_PATTERN.match(pdf_path.name)
        if not match:
            continue
        day = match.group("day")
        shift = match.group("shift")
        reference_path = TESTS_DIRECTORY / f"Target-Etalon-{day}-{shift}.xlsx"
        if reference_path.exists():
            pairs.append(GoldStandardPair(pdf_path, reference_path, day, shift))
    return pairs


GOLD_STANDARD_PAIRS = _gold_standard_pairs()


def _data_cells(workbook_path_or_stream: Path | BytesIO) -> dict[tuple[str, str, str], object]:
    workbook = load_workbook(workbook_path_or_stream, data_only=True)
    cells: dict[tuple[str, str, str], object] = {}
    for sheet in workbook.worksheets:
        if sheet.title == "Rapport_Audit":
            continue
        header_row = next(
            (
                row
                for row in range(1, min(sheet.max_row, 10) + 1)
                if {cell.value for cell in sheet[row]}.issuperset({"Département", "Catégorie", *COLUMNS_TO_COMPARE})
            ),
            None,
        )
        if header_row is None:
            continue
        headers = {cell.value: cell.column for cell in sheet[header_row]}
        if not all(header in headers for header in ("Département", "Catégorie", *COLUMNS_TO_COMPARE)):
            continue
        for row in range(header_row + 1, sheet.max_row + 1):
            department = sheet.cell(row, headers["Département"]).value
            category = sheet.cell(row, headers["Catégorie"]).value
            if not department or not category:
                continue
            for column in COLUMNS_TO_COMPARE:
                value = sheet.cell(row, headers[column]).value
                if column == "Écart (Décompte vs Cible)" and value is None:
                    value = sheet.cell(row, headers["Présences"]).value - sheet.cell(row, headers["Cible"]).value
                cells[(f"{department}/{category}", column, sheet.title)] = value
    return cells


def _recurrence_type(cell_key: tuple[str, str, str], actual: object, expected: object) -> str:
    department, column, _ = cell_key[0].partition("/")[0], cell_key[1], cell_key[2]
    if department == "6e":
        return "Erreurs de détection d'ancres"
    if column == "Présences" and (actual == 0 or expected == 0):
        return "Erreurs de saut de page/date"
    return "Erreurs de déduction de codes"


def _print_audit(results: list[tuple[GoldStandardPair, int, int, list[tuple[str, tuple[str, str, str], object, object]]]]) -> None:
    total_compared = sum(compared for _, compared, _, _ in results)
    total_matches = sum(matches for _, _, matches, _ in results)
    precision = 100 * total_matches / total_compared if total_compared else 0
    print(f"\nPrécision globale : {precision:.1f}% ({total_matches}/{total_compared} cellules)")
    for pair, compared, matches, _ in results:
        score = 100 * matches / compared if compared else 0
        print(f"{pair.day.title()}: {score:.1f}% ({matches}/{compared} cellules)")

    failures = [failure for _, _, _, pair_failures in results for failure in pair_failures]
    if failures:
        print("\nMatrice des écarts :")
        for recurrence, count in sorted(Counter(failure[0] for failure in failures).items()):
            print(f"- {recurrence} ({count})")
            for _, cell_key, actual, expected in (failure for failure in failures if failure[0] == recurrence):
                print(f"  {cell_key[2]} | {cell_key[0]} | {cell_key[1]}: généré={actual!r}, étalon={expected!r}")


@pytest.mark.skipif(not GOLD_STANDARD_PAIRS, reason="Aucune paire PDF/XLSX étalon versionnée dans tests/")
def test_workforce_pipeline_matches_all_gold_standards() -> None:
    """Compare every discovered PDF conversion with its paired gold-standard workbook."""
    results = []
    for pair in GOLD_STANDARD_PAIRS:
        parsed = PDFDocumentParser().parse(pair.pdf_path)
        assert parsed.get("status") == "success", parsed.get("message", "Extraction PDF impossible")
        records = parse_workforce_text(parsed["raw_text"], pair.shift)
        generated_cells = _data_cells(build_workforce_workbook(records, pair.shift))
        reference_cells = _data_cells(pair.reference_path)
        all_cells = sorted(set(generated_cells) | set(reference_cells))
        failures = []
        matches = 0
        for cell_key in all_cells:
            actual = generated_cells.get(cell_key)
            expected = reference_cells.get(cell_key)
            if actual == expected:
                matches += 1
            else:
                failures.append((_recurrence_type(cell_key, actual, expected), cell_key, actual, expected))
        results.append((pair, len(all_cells), matches, failures))

    _print_audit(results)
    failures = [failure for _, _, _, pair_failures in results for failure in pair_failures]
    assert not failures, "Des cellules diffèrent des étalons; consulter la matrice des écarts affichée."