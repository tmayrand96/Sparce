#!/usr/bin/env python
"""
Validation script to verify business rule updates in the generated Excel file.
"""

import sys
from pathlib import Path
from openpyxl import load_workbook


def _data_rows(sheet):
    """Index report rows by their business key."""
    return {
        (row[0], row[1]): row[2:5]
        for row in sheet.iter_rows(min_row=5, values_only=True)
        if row[0] and row[1]
    }


def compare_with_gold_standard(output_path: Path, target_path: Path) -> int:
    """Print every business-value difference between generated and target workbooks."""
    generated = load_workbook(output_path, data_only=True)
    target = load_workbook(target_path, data_only=True)
    differences = 0
    fields = ("Cible", "Présences", "Écart (Décompte vs Cible)")

    for sheet_name in sorted(set(generated.sheetnames) | set(target.sheetnames)):
        if sheet_name == "Rapport_Audit":
            continue
        if sheet_name not in generated.sheetnames:
            print(f"FEUILLE ABSENTE (généré) : {sheet_name}")
            differences += 1
            continue
        if sheet_name not in target.sheetnames:
            print(f"FEUILLE ABSENTE (étalon) : {sheet_name}")
            differences += 1
            continue

        generated_rows = _data_rows(generated[sheet_name])
        target_rows = _data_rows(target[sheet_name])
        for department, category in sorted(set(generated_rows) | set(target_rows)):
            key = (department, category)
            if key not in generated_rows:
                print(f"LIGNE ABSENTE (généré) : {sheet_name} | {department} | {category}")
                differences += 1
                continue
            if key not in target_rows:
                print(f"LIGNE ABSENTE (étalon) : {sheet_name} | {department} | {category}")
                differences += 1
                continue
            for field, extracted, expected in zip(fields, generated_rows[key], target_rows[key]):
                if extracted != expected:
                    print(
                        f"ÉCART : {sheet_name} | Département={department} | "
                        f"Catégorie={category} | {field} | extrait={extracted!r} | étalon={expected!r}"
                    )
                    differences += 1

    print(f"\nTotal des écarts métier : {differences}")
    return differences


def main():
    output_path = Path("tests/output_Rapport-Template-Soir.xlsx")
    target_path = Path("tests/Rapport-Target-Soir.xlsx")
    
    if not output_path.exists():
        print(f"ERROR: Output file not found at {output_path}")
        sys.exit(1)
    if not target_path.exists():
        print(f"ERROR: Gold Standard not found at {target_path}")
        sys.exit(1)
    
    workbook = load_workbook(output_path)
    
    print("\n" + "=" * 80)
    print("WORKBOOK STRUCTURE & VERIFICATION")
    print("=" * 80)
    print(f"\nSheet names: {workbook.sheetnames}")
    
    # Check Audit sheet
    if "Rapport_Audit" in workbook.sheetnames:
        audit_sheet = workbook["Rapport_Audit"]
        print("\n✓ Audit sheet present")
        print(f"  - Title: {audit_sheet['A1'].value}")
        print(f"  - Execution info: {audit_sheet['A2'].value}")
    
    # Check data sheets (all except Audit)
    data_sheets = [name for name in workbook.sheetnames if name != "Rapport_Audit"]
    
    for sheet_name in data_sheets:
        sheet = workbook[sheet_name]
        print(f"\n{'─' * 80}")
        print(f"Sheet: {sheet_name}")
        print(f"{'─' * 80}")
        
        shift = sheet['A1'].value
        date = sheet['A2'].value
        print(f"  Shift: {shift}")
        print(f"  Date: {date}")
        
        # Extract data rows
        rows = []
        for row in sheet.iter_rows(min_row=5, max_row=sheet.max_row, values_only=False):
            if any(cell.value for cell in row):
                dept = row[0].value
                category = row[1].value
                cible = row[2].value
                presences = row[3].value
                ecart_decompte = row[4].value
                
                if dept and category:
                    rows.append({
                        'Département': dept,
                        'Catégorie': category,
                        'Cible': cible,
                        'Présences': presences,
                        'Écart (Déc. vs Cible)': ecart_decompte,
                    })
        
        # Print rows organized by department
        departments_found = {}
        for row in rows:
            dept = row['Département']
            if dept not in departments_found:
                departments_found[dept] = []
            departments_found[dept].append(row)
        
        print("\n  Data by Department:")
        for dept in sorted(departments_found.keys()):
            dept_rows = departments_found[dept]
            print(f"\n    {dept}:")
            for row in dept_rows:
                print(f"      {row['Catégorie']:15s} | Cible: {str(row['Cible']):>2} | "
                          f"Prés: {str(row['Présences']):>2} | Écart: {row['Écart (Déc. vs Cible)']}")
        
        # Verify business rules
        print(f"\n  Verification:")
        has_8e = "8e" in departments_found
        has_7e = "7e" in departments_found
        has_aa = any(row['Catégorie'] == 'AA' for row in rows)
        has_acur_gdl = "ACUR/GDL" in departments_found
        
        print(f"    ✓ Floor 6e mapped to '8e': {has_8e}")
        print(f"    ✓ Floor 7e present: {has_7e}")
        print(f"    ✓ Category 'AA' present: {has_aa}")
        if has_aa:
            aa_rows = [r for r in rows if r['Catégorie'] == 'AA']
            print(f"      - Found {len(aa_rows)} AA entries")
            for dept_name, dept_list in departments_found.items():
                aa_in_dept = [r for r in dept_list if r['Catégorie'] == 'AA']
                if aa_in_dept:
                    print(f"        - {len(aa_in_dept)} in {dept_name}")
        print(f"    ✓ ACUR/GDL department present: {has_acur_gdl}")
    
    print("\n" + "=" * 80)
    print("AUDIT DE PRÉCISION CONTRE L'ÉTALON")
    print("=" * 80)
    return compare_with_gold_standard(output_path, target_path)

if __name__ == "__main__":
    sys.exit(main())
